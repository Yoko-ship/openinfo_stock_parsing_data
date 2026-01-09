from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter



def styling(data_sheet,ai_sheet,ai_text):

    FILE_PATH = "data.xlsx"
    DATA_SHEET = data_sheet
    AI_SHEET = ai_sheet



    GREEN = PatternFill("solid", fgColor="C6EFCE")
    RED = PatternFill("solid", fgColor="FFC7CE")
    GRAY = PatternFill("solid", fgColor="E7E6E6")
    YELLOW = PatternFill("solid", fgColor="FFF2CC")

    BOLD = Font(bold=True)
    ITALIC = Font(italic=True)

    CENTER = Alignment(horizontal="center")
    WRAP_TOP = Alignment(wrap_text=True, vertical="top")



    POSITIVE_COLS = [
        "Чистая Прибыль",
        "Чистая прибыль сум",
        "Чистая выручка",
        "ROE",
        "Рентабельность капитала",
        "Маржа EBIT"
    ]

    DEBT_COLS = [
        "Долгосрочные обязательства, всего",
        "Текущие обязательства, всего",
        "Общие обязательства",
        "Отношение долга к собственному капиталу"
    ]



    wb = load_workbook(FILE_PATH)
    ws = wb[DATA_SHEET]



    header_map = {
        ws.cell(row=1, column=col).value: col
        for col in range(1, ws.max_column + 1)
    }



    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = BOLD
        ws.column_dimensions[get_column_letter(col)].width = 24



    for row in range(2, ws.max_row + 1):
        for col_name, col_idx in header_map.items():
            cell = ws.cell(row=row, column=col_idx)

            if not isinstance(cell.value, (int, float)):
                continue


            if cell.value == 0:
                cell.fill = GRAY
                cell.font = ITALIC
                cell.alignment = CENTER
                continue


            if col_name in POSITIVE_COLS:
                if cell.value > 0:
                    cell.fill = GREEN
                else:
                    cell.fill = RED


            if col_name in DEBT_COLS:
                if cell.value > 1:
                    cell.fill = RED
                elif cell.value < 0.5:
                    cell.fill = GREEN
                else:
                    cell.fill = YELLOW

            cell.alignment = CENTER

    if AI_SHEET in wb.sheetnames:
        ws_ai = wb[AI_SHEET]
        ws_ai.delete_rows(1, ws_ai.max_row)
    else:
        ws_ai = wb.create_sheet(AI_SHEET)

    ws_ai["A1"] = "Инвестиционный анализ"
    ws_ai["A1"].font = Font(bold=True, size=14)

    ws_ai["A3"] = ai_text
    ws_ai["A3"].alignment = WRAP_TOP

    ws_ai.column_dimensions["A"].width = 250
    ws_ai.row_dimensions[3].height = 800


    wb.save(FILE_PATH)

    print("✅ Excel файл успешно стилизован и сохранён")
